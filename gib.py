from selenium.webdriver.support.ui import Select
from selenium import webdriver
import time
from termcolor import cprint
import zipfile
import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl import load_workbook


class Gib(object):

    def __init__(self, user_name, password):
        self.user_name = user_name
        self.password = password
        self.Invoince = []
        self.path = "driver/chromedriver"
        self.options = webdriver.ChromeOptions()
        prefs = {"download.default_directory": "/home/macir/Desktop/DownloadsInvoinces/Downloladeds"}
        self.options.add_experimental_option("prefs",prefs)
        self.options.add_argument("--start-maximized")


    def openGib(self,):
        self.browser = webdriver.Chrome(executable_path=self.path, chrome_options=self.options)
        self.browser.get("https://earsivportal.efatura.gov.tr/intragiris.html")
        self.browser.find_element_by_id("userid").send_keys(self.user_name)
        self.browser.find_element_by_id("password").send_keys(self.password)
        self.browser.find_element_by_css_selector("#formdiv > div:nth-child(4) > div > button").click()
        cprint("Sayfa Yükleniyor....", "blue")
        time.sleep(4)
        self.select = Select(self.browser.find_element_by_id('gen__1006'))
        self.select.select_by_index(1)

    def goModulTaslak(self,):
        #self.first_date = first_date
        #self.last_date = last_date
        cprint("Taslaklar Bölümüne Giriyor...","blue")
        time.sleep(2)
        self.browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/div/div/div/div/div/div[2]/div[2]/div/div/div/div/div[1]/div/div/div/div/div/div/ul/li[2]/a").click()
        self.browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/div/div/div/div/div/div[2]/div[2]/div/div/div/div/div[1]/div/div/div/div/div/div/ul/li[2]/ul/li[2]/a").click()
        self.first_date = (input("Please Enter The Firs Date of The Invoince : "))
        self.last_date = (input("Please Enter The Last Date of The Invoince : "))

        self.browser.find_element_by_id("date-gen__1024").send_keys("  "+self.first_date)
        time.sleep(1)
        self.browser.find_element_by_id("date-gen__1025").send_keys("  " + self.last_date)
        self.browser.find_element_by_id("gen__1026").click()
        time.sleep(5)


    def findbyName(self):
        pass

    def downloadInvoinces(self,):
        try:
            self.total_invoince = self.browser.find_element_by_xpath("//*[@id='gen__1034-pr-td']/span").text[-4:]
            self.total_invoince = int(self.total_invoince)
        except:

            try:
                self.total_invoince = self.browser.find_element_by_xpath("//*[@id='gen__1034-pr-td']/span").text[-3:]
                self.total_invoince = int(self.total_invoince)
            except:

                try:
                    self.total_invoince = self.browser.find_element_by_xpath("//*[@id='gen__1034-pr-td']/span").text[-2:]
                    self.total_invoince = int(self.total_invoince)
                except:

                    try:
                        self.total_invoince = self.browser.find_element_by_xpath("//*[@id='gen__1034-pr-td']/span").text[-1:]
                        self.total_invoince = int(self.total_invoince)
                    except:
                        self.total_invoince = self.browser.find_element_by_xpath("//*[@id='gen__1034-pr-td']/span").text[-5:]
                        self.total_invoince = int(self.total_invoince)
        if self.total_invoince/10 > self.total_invoince//10:
            self.total_page = self.total_invoince//10 + 1
        elif self.total_invoince/10 == self.total_invoince//10:
            self.total_page = self.total_invoince/10



        while self.total_page  >= 1:
            if self.total_page != 1 :
                counter = 1
                counterLag = 0
                while counter <= 10:

                    self.browser.find_element_by_css_selector("#gen__1034-b > tr:nth-child({}) > td.csc-table-select > input[type=checkbox]".format(counter)).click()
                    self.browser.find_element_by_id("gen__1033").click()
                    time.sleep(0.15)
                    if counterLag != 0:
                        self.browser.find_element_by_css_selector("#gen__1034-b > tr:nth-child({}) > td.csc-table-select > input[type=checkbox]".format(counter)).click()
                    elif counterLag == 0 :
                        self.browser.find_element_by_css_selector("#gen__1034-b > tr:nth-child(1) > td.csc-table-select > input[type=checkbox]").click()
                    counter += 1
                    counterLag += 1
                    #time.sleep(1)
                self.browser.find_element_by_css_selector("#gen__1034-div > span.csc-table-paging-btn.csc-table-seek-next").click()
                self.total_page  -= 1
            elif self.total_page == 1 :
                counter = 1
                counterLag = 0
                while counter <= self.total_invoince % 10:

                    self.browser.find_element_by_css_selector(
                        "#gen__1034-b > tr:nth-child({}) > td.csc-table-select > input[type=checkbox]".format(
                            counter)).click()
                    self.browser.find_element_by_id("gen__1033").click()
                    time.sleep(0.15)
                    if counterLag != 0:
                        self.browser.find_element_by_css_selector(
                            "#gen__1034-b > tr:nth-child({}) > td.csc-table-select > input[type=checkbox]".format(
                                counter)).click()
                    elif counterLag == 0:
                        self.browser.find_element_by_css_selector(
                            "#gen__1034-b > tr:nth-child(1) > td.csc-table-select > input[type=checkbox]").click()
                    counter += 1
                    counterLag += 1
                    # time.sleep(1)
                # self.browser.find_element_by_css_selector(
                #     "#gen__1034-div > span.csc-table-paging-btn.csc-table-seek-next").click()
                self.total_page -= 1
        self.browser.close()


    def unZip(self):
        lis_of_zips = os.listdir("Downloladeds")
        for i in lis_of_zips:
            with zipfile.ZipFile("Downloladeds/{}".format(i), 'r') as zip_ref:
                zip_ref.extractall("/home/macir/Desktop/DownloadsInvoinces/Unzipped")



############### Below codes are about parsing an XML file #############################
    def files(self):
        filesinlist = os.listdir("Unzipped")
        for afile in filesinlist:
            if afile[-3:] == "xml":
                #print(afile)
                self.readXml(afile)

    def readXml(self,fileName):
        comBasic = "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}"
        comAggregate = "{urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2}"
        tree = ET.parse('Unzipped/{}'.format(fileName))
        root = tree.getroot()
        IDs =  []
        Taxedlines = []
        SubTotalTax = []
        companyName = []
        #Invoince = []
        for iter in root.iter():

            #print(iter.tag)
            # If it is a company, it finds the name of it
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}Name":
                companyName.append(iter.text)

            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}FirstName":
                name = iter.text
            ##### finds surname ######
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}FamilyName":
                surname = iter.text

        ###### ADDRESS ####
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}StreetName":
                self.adress = iter.text


        #### Takes the values which are named with id #####
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}ID":
                customerid = iter.text
                IDs.append(customerid)
        ## Satırlardaki Kdv Oranlarını Yazar
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}Percent":
                #print(iter.text)
                Taxedlines.append(iter.text)

        # Matrah

            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}TaxAmount":
                SubTotalTax.append(iter.text)
        # Toplam Ücret
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}PayableAmount":
                subtotal = iter.text

        # Toplam Ücret
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}TaxExclusiveAmount":
                subtotal_excluviseved = iter.text
        # Date
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}IssueDate":
                date = iter.text
        # Time
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}IssueTime":
                time = iter.text

        # File Name #
            if iter.tag == "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}UUID":
                xmlFile = iter.text

        try:

            #print(name, surname,self.adress, IDs[6], IDs[0], Taxedlines[1:], subtotal_excluviseved, SubTotalTax[0], subtotal, date, time)

            self.Invoince.append([name, surname,IDs[6],IDs[0],self.adress,   (str(Taxedlines[1:])), subtotal_excluviseved, SubTotalTax[0], subtotal, date, time, xmlFile])

        except:
            #print(companyName[4],self.adress, IDs[6], IDs[0], Taxedlines[1:], subtotal_excluviseved, SubTotalTax[0], subtotal, date, time)

            self.Invoince.append([companyName[4],"",IDs[6],IDs[0],self.adress, str(Taxedlines[1:]), subtotal_excluviseved,   SubTotalTax[0], subtotal, date, time,xmlFile])


    def writeXlsx(self,):
        try:
            book = load_workbook("Faturalar.xlsx")
            sheet = book.active
            for invoince in self.Invoince:
                sheet.append(invoince)
        except:
            book = Workbook()
            sheet = book.active

            header = ["İsim/Unvan",
                      "Soyisim",
                      "Tc",
                      "Fatura No",
                      "Adres",
                      "Kdv Oranı",
                      "Matrah",
                      "Kdv",
                      "Toplam Tutar",
                      "Tarih",
                      "Saat",
                      "Kayıtlı Dosya Adı"
                      ]
            sheet.append(header)
            for invoince in self.Invoince:
                sheet.append(invoince)

        book.save("/home/macir/Desktop/DownloadsInvoinces/Faturalar.xlsx")

#
# a = Gib("username","password")
# # #a.readXml("0a5a08b3-d7a9-4864-9244-30e1a0252bca_f.xml")
# # #a.unZip()
# a.files()
# a.writeXlsx()